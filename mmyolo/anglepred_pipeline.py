import math
import os
import openpyxl
import torch
from angle_prediction.loss_function import angle2code, code2angle
from infer_image import InferImage
import mmcv
from sklearn.cluster import KMeans
import numpy as np
from PIL import Image
from angle_prediction import infer
import argparse

parser = argparse.ArgumentParser()
parser.add_argument("--name", type=str)
parser.add_argument("--Eg", type=str)
args = parser.parse_args()
file_name = args.name + '-' + args.Eg

def angle_infer(image, bbox, model, transform):
    image = Image.fromarray(image)
    cropped_patch = image.crop(bbox)
    angle_pred = infer.test(model, transform, cropped_patch)
    return angle_pred


def average(x, y):
    y, x = max(x, y), min(x, y)
    if y - x > x + 360 - y:
        x = x + 360
    return (x+y) / 2 % 360

# input 2 image paths, as a list or tuple output predicted angle
def images2angle(paths, angle_model, transform, detect_model, match_type='Hungary'):
    '''
    修改该函数，使得在匹配之后，每一个bbox有其对应的角度，并在match list中，给出每一个match的角度预测
    '''
    # in bboxes, we save tow list, which contains bboxes of two images
    bboxes = []
    angle_preds = []
    for path in paths:
        img = mmcv.imread(path)
        img = mmcv.imconvert(img, 'bgr', 'rgb')
        bbox = detect_model.ImagePrediction(img).bboxes
        bbox = bbox.tolist()
        angle_pred = []
        for i, one_bbox in enumerate(bbox):
            angle_pred.append(angle_infer(img, bbox[i], angle_model, transform))
            w = one_bbox[2] - one_bbox[0]
            h = one_bbox[3] - one_bbox[1]
            bbox[i][2] = w
            bbox[i][3] = h
            bbox[i].append(0.0)
        angle_preds.append(angle_pred)
        bboxes.append(bbox)
    # print('bbox:\n', bboxes)
    # get ious from one list of bbox to the other
    angles_det_ans = []
    angles_pred_ans = []
    for i in range(len(angle_preds[0])):
        angles_pred_ans.append(angle_preds[0][i].cpu())
    print("detect pred：", angles_det_ans)
    print("angle, pred: ", angles_pred_ans)
    return angles_det_ans, angles_pred_ans


def sort_(elem):
    return int(elem)


def dataset2angle(path):
    angle_model, transform = infer.init_model_trans()
    detect_model = InferImage()
    conditions = os.listdir(path)
    conditions.sort(key=sort_)
    ans = []
    for j, condition in enumerate(conditions):
        print('condition:', condition)
        paths = os.listdir(os.path.join(path, condition))
        paths.sort(reverse=True)
        for i in range(len(paths)):
            paths[i] = os.path.join(path, condition, paths[i])
        ret_det, ret_angle = images2angle(paths, angle_model, transform, detect_model)
        ans += ret_angle
    print(ans)
    return ans


def twoimages2angle():
    folder = r'D:\PaddleDetection\data'
    paths = os.listdir(folder)
    paths.sort(reverse=True)
    for i in range(len(paths)):
        paths[i] = os.path.join(folder, paths[i])
    ans = images2angle(paths)
    print('angle:', ans)

def restore_i_from_point(point):
    x, y = point[0], point[1]
    radians_i = math.acos(x)
    degrees_i = math.degrees(radians_i)
    return degrees_i

def separate_numbers(data):
    new_data = []
    for i in data:
        i = torch.tensor(i).unsqueeze(0)
        point = angle2code(i)
        new_data.append(np.array(point.squeeze(0).cpu()))
    data = new_data
    data_array = np.array(data)
    kmeans = KMeans(n_clusters=2)
    kmeans.fit(data_array.reshape(-1, 3))
    labels = kmeans.labels_
    x_group = data_array[labels == 0]
    y_group = data_array[labels == 1]
    if len(x_group) < len(y_group):
        x_group, y_group = y_group, x_group
    x_angle = []
    y_angle = []
    for p in x_group:
        out = code2angle(torch.tensor(p).unsqueeze(0).cuda()).float() / torch.pi*180+180
        x_angle.append(out)
    for p in y_group:
        out = code2angle(torch.tensor(p).unsqueeze(0).cuda()).float() / torch.pi*180+180
        y_angle.append(out)
    return x_angle, y_angle


if __name__ == '__main__':
    path = r'D:\wise_transportation\data\2frame_dataset\140-6-9'
    ans = dataset2angle(path)
    # 调用函数进行分离
    x_group, y_group = separate_numbers(ans)
    print("正向角度：", x_group)
    print("负向角度：", y_group)
    F = len(x_group)
    R = len(y_group)
    ratio = R / (F + R)
    print('Forward Number:', F)
    print('Reverse Number:', R)
    print('Wrong-way-cycling prob', ratio)

    # 打开Excel文件，如果文件不存在则创建新文件
    workbook = openpyxl.load_workbook('data.xlsx')
    # 选择或创建一个工作表
    sheet = workbook.active
    # 在最后一行追加数据
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value='angle')
    sheet.cell(row=row, column=2, value=file_name)
    sheet.cell(row=row, column=3, value=F)
    sheet.cell(row=row, column=4, value=R)
    sheet.cell(row=row, column=5, value=ratio)

    # 保存Excel文件
    workbook.save('data.xlsx')