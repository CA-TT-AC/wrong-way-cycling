import math
import os

import numpy as np
import torch

from angle_prediction.loss_function import code2angle, angle2code
from infer_image import InferImage, draw_bbox_on_image
import mmcv
from pipeline_utils import box_iou_rotated
from sklearn.cluster import KMeans
from scipy.optimize import linear_sum_assignment
import numpy as np
import openpyxl
import argparse

parser = argparse.ArgumentParser()
parser.add_argument("--name", type=str)
parser.add_argument("--Eg", type=str)
args = parser.parse_args()
file_name = args.name + '-' + args.Eg
def get_angle(x, y):
    angle = math.atan2(y, x) * 180 / math.pi
    return angle if angle >= 0 else 360 + angle

# input 2 image paths, as a list or tuple output predicted angle
def images2angle(paths, inited_model, match_type='Hungary'):
    bboxes = []
    for path in paths:
        img = mmcv.imread(path)
        img = mmcv.imconvert(img, 'bgr', 'rgb')
        bbox = inited_model.ImagePrediction(img).bboxes
        bbox = bbox.tolist()
        # 可视化
        draw_bbox_on_image(img, bbox, path.split('\\')[-2] + '_' + path.split('\\')[-1])
        for i, one_bbox in enumerate(bbox):
            w = one_bbox[2] - one_bbox[0]
            h = one_bbox[3] - one_bbox[1]
            bbox[i][2] = w
            bbox[i][3] = h
            bbox[i].append(0.0)
        bboxes.append(bbox)

    # print('bbox:\n', bboxes)
    # get ious from one list of bbox to the other
    ious = box_iou_rotated(torch.tensor(bboxes[0]).float(),
                           torch.tensor(bboxes[1]).float()).cpu().numpy()
    # print('iou matrix:\n', ious)
    if len(ious) == 0 or len(ious[0]) == 0:
        return []
    match_list = []
    if match_type == 'Hungary':
        ious[ious > 0.98] = 0
        ious[ious < 0.5] = 0
        # print(ious)
        # 删除全为0的行和列
        nonzero_rows = np.any(ious != 0, axis=1)  # 找到非零行
        nonzero_cols = np.any(ious != 0, axis=0)  # 找到非零列
        ious = ious[nonzero_rows][:, nonzero_cols]  # 使用布尔索引获取非零行和列的子矩阵
        print(ious)
        row_id, col_id = linear_sum_assignment(ious, maximize=True)
        match_list = np.array([row_id, col_id]).transpose()
    else:
        iou_argmax = np.argmax(ious, axis=1)
        iou_max = np.max(ious, axis=1)
        enough_iou_idxs = np.where(iou_max > 0.05)
        for idx in enough_iou_idxs[0]:
            match_list.append((idx, iou_argmax[idx]))
    angles = []
    for i, j in match_list:
        point1 = bboxes[0][i]
        point2 = bboxes[1][j]
        center1 = point1[0] + point1[2] / 2, point1[1] + point1[3] / 2
        center2 = point2[0] + point2[2] / 2, point2[1] + point2[3] / 2
        vec = np.array(center1) - np.array(center2)
        angle = get_angle(vec[0], -vec[1])
        angles.append(angle)
    print(angles)

    return angles

def sort_(elem):
    return int(elem)
def dataset2angle(path):
    conditions = os.listdir(path)
    conditions.sort(key=sort_)
    detect_model = InferImage()
    ans = []
    for j, condition in enumerate(conditions):
        print('condition:', condition)
        paths = os.listdir(os.path.join(path, condition))
        paths.sort(reverse=True)
        for i in range(len(paths)):
            paths[i] = os.path.join(path, condition, paths[i])
        ret = images2angle(paths, detect_model)
        ans += ret
    return ans
def twoimages2angle():
    folder = r'D:\PaddleDetection\data'
    paths = os.listdir(folder)
    paths.sort(reverse=True)
    for i in range(len(paths)):
        paths[i] = os.path.join(folder, paths[i])
    ans = images2angle(paths)
    print('angle:', ans)

def separate_numbers(data):
    new_data = []
    for i in data:
        i = torch.tensor(i).unsqueeze(0)
        point = angle2code(i)
        new_data.append(np.array(point.squeeze(0).cpu()))
    data = new_data
    data_array = np.array(data)
    kmeans = KMeans(n_clusters=2)
    kmeans.fit(data_array.reshape(-1, 3))
    labels = kmeans.labels_
    x_group = data_array[labels == 0]
    y_group = data_array[labels == 1]
    if len(x_group) < len(y_group):
        x_group, y_group = y_group, x_group
    x_angle = []
    y_angle = []
    for p in x_group:
        out = code2angle(torch.tensor(p).unsqueeze(0).cuda()).float() / torch.pi*180+180
        x_angle.append(out)
    for p in y_group:
        out = code2angle(torch.tensor(p).unsqueeze(0).cuda()).float() / torch.pi*180+180
        y_angle.append(out)
    return x_angle, y_angle


if __name__ == '__main__':
    path = r'D:\wise_transportation\data\2frame_dataset\\' + file_name
    ans = dataset2angle(path)
    small = 0
    big = 0
    for x in ans:
        if x < 180:
            small += 1
        else:
            big += 1
    print(small, big)
    # 调用函数进行分离
    x_group, y_group = separate_numbers(ans)
    print("正向角度：", x_group)
    print("负向角度：", y_group)
    F = len(x_group)
    R = len(y_group)
    ratio = R / (F + R)
    print('Forward Number:', F)
    print('Reverse Number:', R)
    print('Wrong-way-cycling prob', R / (F + R))

    # 打开Excel文件，如果文件不存在则创建新文件
    workbook = openpyxl.load_workbook('data.xlsx')
    # 选择或创建一个工作表
    sheet = workbook.active
    # 在最后一行追加数据
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value='detect')
    sheet.cell(row=row, column=2, value=file_name)
    sheet.cell(row=row, column=3, value=F)
    sheet.cell(row=row, column=4, value=R)
    sheet.cell(row=row, column=5, value=ratio)

    # 保存Excel文件
    workbook.save('data.xlsx')

